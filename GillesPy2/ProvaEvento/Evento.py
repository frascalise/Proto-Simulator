import gillespy2
import matplotlib.pyplot as plt
from Module import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def main():

    # Inizializzo gli che conterranno tutte le informazioni lette da chimica.txt
    species = [] 
    frequences = []
    reactions = []

    numTrajectories=1 # Numero lanci da fare
    model = protoZero(species, frequences, reactions)
    results = model.run(number_of_trajectories = numTrajectories)

    # Creo il foglio dove scrivere i dati
    wb = Workbook()
    ws = wb.active

    rowIndex = 1
    columnIndex = 1
    for i in species:
        if i.name != 'dummy':
            cell = ws.cell(row=rowIndex, column=columnIndex, value=i.name)
            redBG = PatternFill(start_color="E97451", end_color="E97451", fill_type="solid")
            cell.fill = redBG

            columnIndex += 1

    columnIndex = 1
    rowIndex = 2

    for index in range(0, numTrajectories):
        trajectory = results[index]

        for i in species:
            if i.name != 'dummy':
                plt.plot(trajectory['time'], trajectory[i.name], label = i.name)

                for yIndex in range(1, len(trajectory[i.name])):
                    ws.cell(row=rowIndex, column=columnIndex, value=trajectory[i.name][yIndex])
                    rowIndex += 1
                
                columnIndex += 1
                rowIndex = 2




    # Salva il workbook su file
    wb.save("output.xlsx")

    plt.legend()
    plt.title("Esempio GillesPy") 
    plt.show()

if __name__ == "__main__":
    main()

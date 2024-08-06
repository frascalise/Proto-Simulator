import os
import matplotlib.pyplot as plt             # type: ignore
from openpyxl.styles import PatternFill     # type: ignore
from openpyxl import Workbook               # type: ignore

from modules.Module import *
from modules.ReadParams import *


def outputData(species, reactions, catalysis, events, frequences, genCounter):
    print("********** GENERAZIONE ", genCounter + 1, "********** \n")
    print("\n########## SPECIE ##########\n")
    for i in species:
        print(i)
    print("\n=================================")
    print("\n########## REAZIONI ##########\n")
    for i in reactions:
        print(i)
    print("\n=================================")
    print("\n########## FREQUENZE ##########\n")
    for i in frequences:
        print(i)
    print("\n=================================")
    print("\n########## CATALISI ##########\n")
    for i in catalysis:
        print(i, ":", catalysis[i])
    print("\n=================================")
    print("\n########## EVENTI ##########\n")
    for i in events:
        print(i)
    print("\n=================================")
    print("\n")


def initialize(species, frequences, reactions, catalysis, events):
    species.clear()
    frequences.clear()
    reactions.clear()
    catalysis.clear()
    events.clear()


def addEmptyRows(ws):
    insertRow = False
    for row in range(3, ws.max_row + 1):
        if insertRow:
            insertRow = False
            continue
        # Se la cella precedente è uno zero e la cella corrente non è uno zero
        if ws.cell(row=row, column=1).value == 0 and insertRow == False:
            # Inserisci una riga tra la cella precedente e la cella corrente
            ws.insert_rows(row)
            insertRow = True


def fixExcel(ws, speciesColumn, catalysis):
    lipidName = list(catalysis.keys())[0]
    #print("Il nome del lipide e': ", lipidName)

    # Assegno i valori di "TIME" e "tempo" alle righe vuote appena create
    for row in range(2, ws.max_row + 1):
        if all(ws.cell(row=row, column=cell).value is None for cell in range(2, ws.max_column + 1)):
            for column in range(0, ws.max_column): # ws.max_column + 1 è il numero di colonne che conto
                #print(speciesColumn[column])
                if speciesColumn[column] == "TIME":
                    key = [k for k, v in speciesColumn.items() if v == "tempo"][0]
                    ws.cell(row=row, column=column+1, value=ws.cell(row=row+1, column=key+1).value)
                    ws.cell(row=row, column=key+1, value=ws.cell(row=row+1, column=key+1).value)

    # Assegno i valori del lipide alle righe vuote appena create, ovvero il doppio del valore della riga successiva
    lipidKey = [k for k, v in speciesColumn.items() if v == lipidName][0] + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=lipidKey).value is None:
            ws.cell(row=row, column=lipidKey).value = ws.cell(row=row+1, column=lipidKey).value*2
    
    # Assegno i valori delle altre specie alle righe vuote appena create, ovvero il valore della riga successiva/0.35
    for i in range(2, len(speciesColumn.items())-2):
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=i+1).value is None:
                ws.cell(row=row, column=i+1).value = int(ws.cell(row=row+1, column=i+1).value/0.35)

    # Assegno i valori della riga vuota nella colonna "ABSOLUTE TIME"
    ws.cell(row=2, column=2).value = ws.cell(row=2, column=1).value
    for column in range(0, ws.max_column):
        if speciesColumn[column] == "ABSOLUTE TIME":
            for row in range(3, ws.max_row + 1):
                absoluteTimeAdd = ws.cell(row=row, column=1).value - ws.cell(row=row-1, column=1).value
                if absoluteTimeAdd < 0:
                    ws.cell(row=row, column=column+1).value = 0 + ws.cell(row=row-1, column=column+1).value
                else:
                    ws.cell(row=row, column=column+1).value = absoluteTimeAdd + ws.cell(row=row-1, column=column+1).value

def addAbsoluteTime(ws):
    # Aggiungi una colonna tra la 1 e la 2 al file excel
    ws.insert_cols(2)
    ws.cell(row=1, column=2, value="ABSOLUTE TIME")
    # Inizializza la colonna "ABSOLUTE TIME" a 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.value = 0
    
    # Aggiorna il valore della prima cella di "ABSOLUTE TIME" con il valore della prima cella della prima colonna
    ws.cell(row=2, column=2, value=ws.cell(row=2, column=1).value)

    # Aggiorna i valori della colonna "ABSOLUTE TIME"
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row - 1, column=2).value < ws.cell(row=row, column=1).value:
            #ws.cell(row=row, column=2).value = ws.cell(row=row, column=1).value
            ws.cell(row=row, column=2).value = 0
        else:
            #ws.cell(row=row, column=2).value = ws.cell(row=row - 1, column=2).value + abs((ws.cell(row=row, column=1).value - ws.cell(row=row - 1, column=1).value))
            ws.cell(row=row, column=2).value = 0


def colorCells(ws):
    # Evidenzia in giallo l'inizio di una nuova protocellula
    insertRow = False
    for row in range(3, ws.max_row + 1):
        if insertRow:
            insertRow = False
            continue
        # Se la cella precedente è uno zero e la cella corrente non è uno zero
        if ws.cell(row=row, column=1).value == 0 and insertRow == False:
            # Imposta lo sfondo giallo per la nuova riga
            for cell in ws[row]:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            insertRow = True

    # Evidenzia in rosso la prima riga con gli indicatori delle colonne
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="8493B0", end_color="8493B0", fill_type="solid")

# Rimuove l'ultima generazione in modo tale da sistemare il bug dell'ultima generazione non stampata
def removeLastGen(ws):
    # Trova l'ultima riga con lo sfondo giallo
    last_yellow_row = None
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=1).fill.start_color.rgb == "00FFFF00":
            last_yellow_row = row
            break
    
    # Rimuovi le righe dal fondo fino all'ultima riga con lo sfondo giallo
    if last_yellow_row is not None:
        ws.delete_rows(last_yellow_row + 1, ws.max_row)


def findDivision(ws, sintesiList, speciesColumn, catalysis):
    
    lipideColumn = None

    for key, value in speciesColumn.items():
        if value == list(catalysis.keys())[0]:
            lipideColumn = key
            break
    
    # print("Lipide column: ", lipideColumn)
    # print("Lipide name: ", list(catalysis.keys())[0])

    if lipideColumn is not None:
        for row in range(3, ws.max_row + 1):
            lipideValue = ws.cell(row=row, column=lipideColumn+1).value
            previousLipideValue = ws.cell(row=row-1, column=lipideColumn+1).value
            if lipideValue is not None and previousLipideValue is not None and int(lipideValue) < int(previousLipideValue):
                sintesiList.append(row)
    else:
        print("Lipide column not found in speciesColumn.")

    # print("Righe massime: ", ws.max_row)
    # input("Premi invio per continuare...")

def fixLipid(ws, sintesiList, speciesColumn, catalysis):
    lipideColumn = None
    # Trova la colonna del lipide
    for key, value in speciesColumn.items():
        if value == list(catalysis.keys())[0]:
            lipideColumn = key
            break
    
    for i in sintesiList:
        ws.cell(row=i-1, column=lipideColumn+1).value = ws.cell(row=i, column=lipideColumn+1).value*2

def main():

    # Inizializzo i parametri che verranno poi letti in params.txt
    INPUT_FILE = readInput()
    OUTPUT_FILE = readOutput()
    TIME = readTime()
    POINTS = readPoints()
#    TRAJECTORIES = readTrajectories()
    TRAJECTORIES = 1
    COEFF = readCoeff() # al posto di VOL ci mettiam (COEFF*LIPIDE)
    GENERATIONS = readGenerations() + 1
    MAX_LIPID = readMaxLipid()
    
    SINTESI_FILE = readSynthesis()
    'output/sintesi.xlsx'

    oldFile = INPUT_FILE    # Salva il file di input originale

    print("INPUT: ", INPUT_FILE)
    print("OUTPUT: ", OUTPUT_FILE)
    print("TIME: ", TIME)
    print("POINTS: ", POINTS)
    print("TRAJECTORIES: ", TRAJECTORIES)
    print("COEFF: ", COEFF)
    print("GENERATIONS: ", GENERATIONS)
    print("\n")

    species = [] 
    frequences = []
    reactions = []
    catalysis = {}
    events = []  

    # Inizializzo le liste e dizionari che conterranno tutte le informazioni lette da chimica.txt
    initialize(species, frequences, reactions, catalysis, events)
    model = protoZero(INPUT_FILE, TIME, POINTS, COEFF, MAX_LIPID, species, frequences, reactions, catalysis, events)

    # Creo il foglio dove scrivere i dati
    wb = Workbook()
    ws = wb.active

    speciesColumn = {}  # Dizionario che mappa la colonna con il nome della specie

    # Aggiungi la colonna TIME al foglio di lavoro
    ws.cell(row=1, column=1, value="TIME")
    speciesColumn[0] = "TIME"


    orderedSpecies = []
    for k in species:
        orderedSpecies.append(k.name)
    
    orderedSpecies.sort()
    orderedSpeciesColumn = {}
    
    rowIndex = 1
    columnIndex = 2
    for j in range(0, TRAJECTORIES):
        for i in orderedSpecies:
            cell = ws.cell(row=rowIndex, column=columnIndex, value=i)
            orderedSpeciesColumn[columnIndex-1] = i
            redBG = PatternFill(start_color="E97451", end_color="E97451", fill_type="solid")
            cell.fill = redBG

            columnIndex += 1

    speciesColumn.update(orderedSpeciesColumn)
    #print(speciesColumn)

    continueRow = 2 # Indice della riga da cui continuare a scrivere i dati
    continueTime = 0
    stopSimulation = False
    
    for genCounter in range(0, GENERATIONS):
        initialize(species, frequences, reactions, catalysis, events)
        stopGeneration = False
        dummyValues = {} # Dizionario temporanea per salvare i valori delle specie

        model = protoZero(INPUT_FILE, TIME, POINTS, COEFF, MAX_LIPID, species, frequences, reactions, catalysis, events)
        
        os.system('cls' if os.name == 'nt' else 'clear')
        totalGenerations = genCounter + 1
        # Non stampa la generazione non esistente
        if genCounter < GENERATIONS - 1:
            outputData(species, reactions, catalysis, events, frequences, genCounter)

        results = model.run(number_of_trajectories = TRAJECTORIES)
        
        for index in range(0, TRAJECTORIES):
            trajectory = results[index]
            #print("Trajectory: ", trajectory)

            for i in range(0, len(trajectory["time"])): # Scorro per tutta la lunghezza dei dati
                for j in range(0, len(trajectory)):     # Scorro per ogni specie, quindi inserimento per riga 

                    if speciesColumn[j] == "tempo":
                        ws.cell(row=i+continueRow, column=j+1, value=int(continueTime))

                        # Se il tempo cambia, vuol dire che la protocellula si e' divisa e la generazione finisce
                        if trajectory[j][i] != trajectory[j][i-1] and i != 0:       
                            #print("STOP GENERATION: ", stopGeneration, "\n")
                            #print("STOP SIMULATION: ", stopSimulation, "\n")
                            #print("continueTime + trajectory[j][i] =",continueTime, "+", trajectory[j][i])
                            continueTime = trajectory[j][i]
                            stopGeneration = True
                            stopSimulation = False     
                    else:
                        ws.cell(row=i+continueRow, column=j+1, value=int(trajectory[j][i]))
                
                if stopGeneration:
                    continueRow += i

                    # Salva i valori delle specie in dummyValues
                    for j in range(1, len(trajectory)-1):
                        dummyValues[speciesColumn[j]] = int(trajectory[j][i])
                    
                    #print("DummyValues: ", dummyValues)

                    # Cancella il file "input/dummyChimica.txt" se esiste
                    if os.path.exists("input/KBVNRDL1Qp_Chimica.txt"):
                        os.remove("input/KBVNRDL1Qp_Chimica.txt")

                    # Crea un nuovo file di testo chiamato KBVNRDL1Qp_Chimica nella cartella input in scrittura, conterra' le nuove quantita' delle specie
                    # e le reazioni chimiche originali
                    with open("input/KBVNRDL1Qp_Chimica.txt", "w") as file:
                        INPUT_FILE = "input/KBVNRDL1Qp_Chimica.txt"
                        #print(catalysis)                        
                        
                        #   Scrive le nuove quantita' delle specie nel file dummyChimica.txt
                        for k in range(0, len(dummyValues)):
                            speciesFile = list(catalysis.keys())[k] + "\t" + str(dummyValues[list(catalysis.keys())[k]]) + "\t" + catalysis[list(catalysis.keys())[k]] + "\n"
                            file.write(speciesFile)
                        
                        file.write("\n")

                        # Apro in lettura la chimica originale per salvarmi le reazioni chimiche
                        with open(str(oldFile), "r") as of:
                            content = of.read()
                            empty_line_index = content.index("\n\n")

                            # scrivo tutto quello che trova dopo la riga vuota in una stringa
                            new_content = content[empty_line_index+2:]

                        # Scrivo le reazioni chimiche nel file dummyChimica.txt
                        file.write(new_content)
                    
                    break
                else:
                    stopSimulation = True
        
        if stopSimulation:
            break


    os.system('cls' if os.name == 'nt' else 'clear')
    
    # Cancella il file "input/KBVNRDL1Qp_Chimica.txt" se esiste
    if os.path.exists("input/KBVNRDL1Qp_Chimica.txt"):
        os.remove("input/KBVNRDL1Qp_Chimica.txt")

#** ======= MANUTENZIONE FOGLIO EXCEL =======

    # Aggiungo le righe vuote dopo ogni divisione 
    addEmptyRows(ws)
    
    # Aggiorna il dizionario speciesColumn
    newSpeciesColumn = {}
    for key, value in speciesColumn.items():
        newSpeciesColumn[key+1] = value
    newSpeciesColumn[0] = "TIME"
    newSpeciesColumn[1] = "ABSOLUTE TIME"
    speciesColumn = newSpeciesColumn
    # Ordina speciesColumn per ordine di chiave
    speciesColumn = dict(sorted(speciesColumn.items(), key=lambda x: x[0]))
    #print(speciesColumn)

    # Aggiungi una colonna tra la 1 e la 2 al file excel contenente i valori di "ABSOLUTE TIME"
    ws.insert_cols(2)
    ws.cell(row=1, column=2, value="ABSOLUTE TIME")

    # Sistema le righe vuote e i valori delle specie
    fixExcel(ws, speciesColumn, catalysis)
    colorCells(ws)

    sintesiList = []
    findDivision(ws, sintesiList, speciesColumn, catalysis)
    
    # Se le generazioni vengono fatte tutte quante allora rimuovo l'ultima generazione
    if totalGenerations == GENERATIONS:
        removeLastGen(ws)
        moreGenerations = True
        sintesiList.pop()
    else:
        # Colora l'ultima riga del file excel in arancione
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        moreGenerations = False
        sintesiList.append(last_row)

    fixLipid(ws, sintesiList, speciesColumn, catalysis)

    # Salva il file excel coi risultati
    wb.save(OUTPUT_FILE)

#** ======= FILE SINTESI =======
    # Crea un nuovo file excel per la sintesi
    wb2 = Workbook()
    ws2 = wb2.active
    
    # Copia la prima riga del file excel wb nel file excel wb2
    for column in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=column).value
        ws2.cell(row=1, column=column, value=value)
    
    # Copia le righe contenute in sintesiList nel nuovo file excel wb2
    for row_number in sintesiList:
        row_values = []
        for column in range(1, ws.max_column + 1):
            value = ws.cell(row=row_number, column=column).value
            row_values.append(value)
        ws2.append(row_values)

    # Salva il file excel coi risultati della sintesi
    wb2.save(SINTESI_FILE)
    
    # Se vengono fatte piu generazioni rispetto a quelle richieste sistemo il numero di generazioni totali
    if moreGenerations:
        print("\n########## SIMULAZIONE TERMINATA CON ", totalGenerations - 1, "GENERAZIONI ##########\n")
    else:
        print("\n########## SIMULAZIONE TERMINATA CON ", totalGenerations, "GENERAZIONI ##########\n")
    
    print("Il file excel con i risultati si trova in: ", OUTPUT_FILE)
    print("Il file excel con i risultati della sintesi si trova in: ", SINTESI_FILE)
    # print("Righe di sintesi: ", sintesiList)
    # print("Righe totali: ", len(sintesiList))
    print("\n##############################################################\n")
    #quotes()
    

if __name__ == "__main__":
    main()

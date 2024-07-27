import os
import platform
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Color

DISTANZA_ANGOLARE = "output/DistanzaAngolare.xlsx"
ANGULAR_PARAMS = "input/AngularParams.txt"

def read_params(file_path):
    params = {}
    with open(file_path, 'r') as file:
        for line in file:
            key, value = line.strip().split()
            params[key] = value
    return params

def write_params(file_path, params):
    with open(file_path, 'w') as file:
        for key, value in params.items():
            file.write(f"{key}\t{value}\n")

def run_script(n, params_path):
    params = read_params(params_path)
    
    # Salvare i valori originali
    original_output_file = params['OUTPUT']
    original_synthesis_file = params['SYNTHESIS']
    
    base_output_file = original_output_file.replace('.xlsx', '')
    base_synthesis_file = original_synthesis_file.replace('.xlsx', '')

    synthesis_files = []
    
    for i in range(1, n + 1):
        output_file = f"{base_output_file}_Sim{i}.xlsx"
        params['OUTPUT'] = output_file
        synthesis_file = f"{base_synthesis_file}_Sim{i}.xlsx"
        params['SYNTHESIS'] = synthesis_file
        synthesis_files.append(f"output/{synthesis_file}")
        write_params(params_path, params)

        print(f"Simulazione {i} di {n} in corso...")
        subprocess.run(['python3', 'Main.py'], capture_output=True, text=True)
        print(f"Simulazione {i} di {n} completata")
        print("\n-------------------------------------------\n")
    
    # Ripristinare i valori originali
    params['OUTPUT'] = original_output_file
    params['SYNTHESIS'] = original_synthesis_file
    write_params(params_path, params)
    
    return synthesis_files

# Crea il file per la distanza angolare (il primo foglio)
def angular_distance_data(synthesis_files):
    # Inizializza il nuovo workbook
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "Dati Simulazione"
    
    # Flag per indicare se la riga di intestazione è già stata copiata
    header_copied = False

    for i, file in enumerate(synthesis_files, start=1):
        wb = load_workbook(file)
        ws = wb.active
        
        if not header_copied:
            # Crea l'intestazione
            ws_new.append(["Simulazione", "Generazioni"] + [cell.value for cell in ws[1]])
            header_copied = True
        
        # Prendi l'ultima riga del foglio e aggiungi il numero di generazioni
        last_row = [f"Simulazione{i}"] + [ws.max_row - 1] + [cell.value for cell in ws[ws.max_row]]
        ws_new.append(last_row)
    
    # Formatta il foglio di lavoro
    format_sheet(ws_new)
    
    # Salva il nuovo file Excel
    wb_new.save(DISTANZA_ANGOLARE)

    # Scrive il numero di generazioni nel file AngularParams.txt
    write_angular_params(len(synthesis_files))

# Funzione per formattare il foglio di lavoro
def format_sheet(ws):
    # Definisci i riempimenti
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_blue = PatternFill(start_color="4287f5", end_color="4287f5", fill_type="solid")
    fill_light_red = PatternFill(start_color="ffcccb", end_color="ffcccb", fill_type="solid")
    
    # Inizialmente tutto bianco
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.fill = fill_white

    # Intestazione in blu
    for cell in ws[1]:
        cell.fill = fill_blue

    # Prima colonna in blu
    for row in ws.iter_rows(min_col=1, max_col=1, min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = fill_blue
    
    # Seconda colonna in rosso chiaro
    for row in ws.iter_rows(min_col=2, max_col=2, min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = fill_light_red

# Scrivo i parametri del file AngularParams.txt
def write_angular_params(generations):
    with open(ANGULAR_PARAMS, 'w') as file:
        file.write(f"GENERATIONS\t{generations}\n")
        file.write("SPECIES\tA\n")

def write_total_simulations(file_path, total_simulations):
    with open(file_path, 'a') as file:
        file.write(f"TOTAL_SIM\t{total_simulations}\n")

def play_beep():
    if platform.system() == 'Windows':
        import winsound
        winsound.Beep(1000, 500)  # Frequenza di 1000 Hz per 500 ms
    elif platform.system() == 'Darwin':  # macOS
        import subprocess
        subprocess.run(['afplay', '-t', '0.5', '/System/Library/Sounds/Hero.aiff'])
        subprocess.run(['afplay', '-t', '0.5', '/System/Library/Sounds/Submarine.aiff'])
        subprocess.run(['afplay', '-t', '0.5', '/System/Library/Sounds/Morse.aiff'])
        subprocess.run(['afplay', '-t', '0.5', '/System/Library/Sounds/Purr.aiff'])
    else:
        print("Sistema operativo non supportato per la riproduzione del suono.")

if __name__ == "__main__":
    # ** ESEGUO N VOLTE LA SIMULAZIONE **
    params_path = 'input/params.txt'
    n = int(input("Quante volte vuoi eseguire main.py? "))
    subprocess.run('cls' if os.name == 'nt' else 'clear', shell=True)
    synthesis_files = run_script(n, params_path)

    # ** CALCOLO LA DISTANZA ANGOLARE **
    angular_distance_data(synthesis_files)
    print(f"Dati delle simulazioni salvati in {DISTANZA_ANGOLARE}\n")

    # ** SCRIVO IL NUMERO TOTALE DI SIMULAZIONI IN ANGULAR_PARAMS **
    write_total_simulations(ANGULAR_PARAMS, n)

    play_beep()

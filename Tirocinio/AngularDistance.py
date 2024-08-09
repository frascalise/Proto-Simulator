import os
import math
from openpyxl import load_workbook, Workbook        # type: ignore
from openpyxl.styles import PatternFill             # type: ignore
from modules.ReadParams import readSynthesis  # Importa la funzione readSynthesis
from modules.ReadParams import readTotalSim  # Importa la funzione readTotalSim

DISTANZA_ANGOLARE = "output/DistanzaAngolare.xlsx"
ANGULAR_PARAMS = "input/AngularParams.txt"
SINTESI = readSynthesis()
TOTAL_SIM = readTotalSim()

def read_species_and_generations(file_path):
    species = []
    generations = 0
    with open(file_path, 'r') as file:
        for line in file:
            key, value = line.strip().split()
            if key == "SPECIES":
                species = value.split(',')
            elif key == "GENERATIONS":
                generations = int(value)
    return species, generations

def load_synthesis_files():
    synthesis_data = {}
    base_name = os.path.splitext(SINTESI)[0]
    for i in range(1, TOTAL_SIM + 1):
        synthesis_file = f"{base_name}_Sim{i}.xlsx"
        if os.path.exists(synthesis_file):
            wb = load_workbook(synthesis_file)
            ws = wb.active
            synthesis_data[f"Sim{i}"] = [row for row in ws.iter_rows(values_only=True)]
    return synthesis_data

def remove_unwanted_columns(data, unwanted_columns):
    headers = data[0]
    indices_to_remove = [i for i, header in enumerate(headers) if header in unwanted_columns]
    cleaned_data = []
    for row in data:
        cleaned_row = [cell for i, cell in enumerate(row) if i not in indices_to_remove]
        cleaned_data.append(cleaned_row)
    return cleaned_data

def filter_data_by_species(data, species):
    headers = data[0]
    species_indices = [headers.index(spec) for spec in species if spec in headers]
    filtered_data = []
    for row in data:
        filtered_row = [row[i] for i in species_indices]
        filtered_data.append(filtered_row)
    return filtered_data

def create_generation_sheets(wb, synthesis_data, num_generations, species):
    blue_fill = PatternFill(start_color="4287f5", end_color="4287f5", fill_type="solid")
    unwanted_columns = ["TIME", "ABSOLUTE TIME"]
    
    for gen in range(num_generations):
        ws = wb.create_sheet(f"Dati Gen{gen + 1}")
        
        # Aggiungi intestazioni (Simulazione + intestazioni del file di sintesi)
        headers = synthesis_data[list(synthesis_data.keys())[0]][0]
        filtered_headers = [header for header in headers if header not in unwanted_columns]
        ws.cell(row=1, column=1, value="Simulazione")
        for col_num, header in enumerate(filtered_headers, start=2):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = blue_fill  # Colora la prima riga di blu
        
        # Aggiungi dati di ogni simulazione per la generazione corrente
        row_num = 2
        for sim_index, (sim_key, data) in enumerate(synthesis_data.items(), start=1):
            cleaned_data = remove_unwanted_columns(data, unwanted_columns)
            if len(cleaned_data) > gen + 1:  # Controlla se la generazione esiste nel file di sintesi
                ws.cell(row=row_num, column=1, value=sim_key).fill = blue_fill  # Colora la prima colonna di blu
                for col_num, cell_value in enumerate(cleaned_data[gen + 1], start=2):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
                row_num += 1

#** Calcola la distanza angolare tra le simulazioni
def calculate_angular_distance(ws, species):
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = rows[1:]
    
    # Trova gli indici delle specie nelle intestazioni
    species_indices = {header: index for index, header in enumerate(headers) if header in species}
    
    # Prepara un foglio per salvare le distanze angolari
    distance_sheet = ws.parent.create_sheet(f"Distanza Angolare {ws.title}")
    
    # Intestazioni della matrice delle distanze
    distance_sheet.cell(row=1, column=1, value="Index")
    for i in range(2, len(data) + 2):
        distance_sheet.cell(row=1, column=i, value=rows[i-1][0])
        distance_sheet.cell(row=i, column=1, value=rows[i-1][0])
    
    # Colora la prima riga e la prima colonna di blu
    blue_fill = PatternFill(start_color="4287f5", end_color="4287f5", fill_type="solid")
    for cell in distance_sheet[1]:
        cell.fill = blue_fill
    for row in distance_sheet.iter_rows(min_col=1, max_col=1):
        for cell in row:
            cell.fill = blue_fill
    
    # Calcolo della distanza angolare (angolo tra i vettori)
    for i in range(len(data)):
        sim1 = [data[i][species_indices[spec]] for spec in species if spec in species_indices]
        distance_sheet.cell(row=i + 2, column=i + 2, value=0)  # Diagonale principale = 0 (distanza con se stesso)
        
        for j in range(i + 1, len(data)):
            if i == j:
                angle = 0  # Distanza angolare tra la stessa simulazione
            else:
                sim2 = [data[j][species_indices[spec]] for spec in species if spec in species_indices]
                
                # Calcola il coseno dell'angolo tra i vettori
                dot_product = sum(a * b for a, b in zip(sim1, sim2))
                magnitude1 = math.sqrt(sum(a**2 for a in sim1))
                magnitude2 = math.sqrt(sum(b**2 for b in sim2))
                
                # Calcola l'angolo (distanza angolare) in gradi
                if magnitude1 != 0 and magnitude2 != 0:
                    cosine_angle = dot_product / (magnitude1 * magnitude2)
                    # Corregge il valore del coseno per rimanere entro il dominio valido
                    cosine_angle = max(min(cosine_angle, 1.0), -1.0)
                    angle = math.degrees(math.acos(cosine_angle))
                else:
                    angle = "Non Valida"  # Gestione del caso di divisione per zero
            
            distance_sheet.cell(row=i + 2, column=j + 2, value=angle)
            distance_sheet.cell(row=j + 2, column=i + 2, value=angle)

#** Legge il numero totale di simulazioni dal file di parametri
def read_total_simulations(file_path):
    total_sim = 0
    with open(file_path, 'r') as file:
        for line in file:
            key, value = line.strip().split()
            if key == "TOTAL_SIM":
                total_sim = int(value)
    return total_sim

#** Aggiorno e riordino le matrici delle distanze angolari, aggiungendo le simulazioni morte
def update_and_reorder_distance_matrices(wb, total_simulations):
    for sheet in wb.sheetnames:
        if sheet.startswith("Distanza Angolare Dati Gen"):
            ws = wb[sheet]
            
            # Leggi le intestazioni esistenti
            existing_headers = [cell.value for cell in ws[1]]
            existing_simulations = existing_headers[1:]  # Le intestazioni delle colonne sono dalla colonna 2 in poi

            # Trova tutte le simulazioni
            all_simulations = [f"Sim{i}" for i in range(1, total_simulations + 1)]

            # Aggiungi righe e colonne mancanti
            new_headers = existing_headers[:]
            for sim in all_simulations:
                if sim not in existing_simulations:

                    # Aggiungi colonna
                    col_idx = all_simulations.index(sim) + 2
                    ws.insert_cols(col_idx)
                    ws.cell(row=1, column=col_idx, value=sim)
                    new_headers.append(sim)
                    
                    # Aggiungi riga
                    row_idx = all_simulations.index(sim) + 2
                    ws.insert_rows(row_idx)
                    ws.cell(row=row_idx, column=1, value=sim)
    
            # Colora la prima riga e la prima colonna di blu
            blue_fill = PatternFill(start_color="4287f5", end_color="4287f5", fill_type="solid")
            for cell in ws[1]:
                cell.fill = blue_fill
            for row in ws.iter_rows(min_col=1, max_col=1):
                for cell in row:
                    cell.fill = blue_fill

#** Aggiorno i dati dei fogli excel Dati GenX con i nomi delle simulazioni morte
def update_datiGen(wb, total_simulations):
    rowNames = []
    for i in range(1, total_simulations + 1):
        rowNames.append(f"Sim{i}")
    
    for sheet in wb.sheetnames:
        if sheet.startswith("Dati Gen"):
            ws = wb[sheet]
            for i in range(1, total_simulations + 1):
                if ws.cell(row=i + 1, column=1).value != rowNames[i - 1]:
                    ws.insert_rows(i + 1)
                    ws.cell(row=i + 1, column=1, value=rowNames[i - 1])

#** Sistemo il foglio excel Sintesi
def blankSintesiSheet(ws, num_generations, total_simulations):
#* Prima tabella coi dati del foglio excel Sintesi
    # Scrivo la prima riga del foglio excel Sintesi
    ws.cell(row=1, column=1, value="Angoli")
    for i in range(1, num_generations + 1):
        ws.cell(row=1, column=i + 1, value=f"Gen{i}")
        ws.cell(row=1, column=i + 1).fill = PatternFill(start_color="4287f5", end_color="4287f5", fill_type="solid")
    
    # Scrivo i tipi di valori delle varie righe
    ws.cell(row=2, column=1, value="Min")
    ws.cell(row=3, column=1, value="Media")
    ws.cell(row=4, column=1, value="Mediana")
    ws.cell(row=5, column=1, value="Max")
    ws.cell(row=6, column=1, value="Dev. Std.")
    ws.cell(row=7, column=1, value="Sim Vive")
    # La riga 8 è vuota
#* Seconda tabella coi tempi del foglio excel Sintesi
    ws.cell(row=9, column=1, value="Tempo")
    for i in range(1, num_generations + 1):
        ws.cell(row=9, column=i + 1, value=f"Gen{i}")
        ws.cell(row=9, column=i + 1).fill = PatternFill(start_color="4287f5", end_color="4287f5", fill_type="solid")
    
    for i in range(1, total_simulations + 1):
        ws.cell(row=9 + i, column=1, value=f"Sim{i}")
        ws.cell(row=9 + i, column=1).fill = PatternFill(start_color="4287f5", end_color="4287f5", fill_type="solid")
    
    # Coloro le righe 2-7 di blu
    for i in range(2, 8):
        ws.cell(row=i, column=1).fill = PatternFill(start_color="4287f5", end_color="4287f5", fill_type="solid")

    # Coloro le rimanenti celle di blu
    ws.cell(row=1, column=1).fill = PatternFill(start_color="4287f5", end_color="4287f5", fill_type="solid")
    ws.cell(row=9, column=1).fill = PatternFill(start_color="4287f5", end_color="4287f5", fill_type="solid")

#** Calcolo della mediana
def calculate_median(values):
    sorted_values = sorted(values)
    n = len(sorted_values)
    midpoint = n // 2
    
    if n % 2 == 0:  # Se pari, la mediana è la media dei due valori centrali
        return (sorted_values[midpoint - 1] + sorted_values[midpoint]) / 2
    else:  # Se dispari, la mediana è il valore centrale
        return sorted_values[midpoint]

#** Calcolo della deviazione standard
def calculate_standard_deviation(values, mean):
    variance = sum((x - mean) ** 2 for x in values) / len(values)
    return math.sqrt(variance)

#** Creo il foglio excel Sintesi
def sintesiSheet(wb, total_simulations, num_generations):
    ws = wb.create_sheet("Sintesi")
    wb._sheets.remove(ws)
    wb._sheets.insert(1, ws)

    blankSintesiSheet(ws, num_generations, total_simulations)

#*  Trascrivo i tempi delle varie simulazioni (seconda tabella del foglio excel Sintesi)
    for i in range(1, num_generations + 1):
        wsDatiGen = wb[f"Dati Gen{i}"]  # Apro i vari fogli "Dati Gen" per prendere i tempi
        
        for j in range(1, wsDatiGen.max_column + 1):    # Cerco la colonna "tempo" nel foglio Dati Gen
            if wsDatiGen.cell(row=1, column=j).value == "tempo":
                for k in range(1, wsDatiGen.max_row):   # Quando trovo la colonna "tempo" trascrivo i tempi
                    ws.cell(row=9+k, column=i+1).value = wsDatiGen.cell(row=k+1, column=j).value

#*  Trascrivo i dati delle varie distanze angolari (prima tabella del foglio excel Sintesi)
    rows, cols = total_simulations, total_simulations   # Righe e colonne della matrice delle distanze angolari
    for i in range(1, num_generations + 1): 
        wsDistanzaAngolare = wb[f"Distanza Angolare Dati Gen{i}"]   # Apro i vari fogli "Distanza Angolare Dati Gen" per prendere i dati
        matrix = [[None for _ in range(cols)] for _ in range(rows)] # Inizializzo la matrice delle distanze angolari nulla
        
        for j in range(1, total_simulations + 1):
            for k in range(1, total_simulations + 1):
                if j != k:
                    matrix[j-2][k-2] = wsDistanzaAngolare.cell(row=j+1, column=k+1).value

        # Estrazione dei valori non nulli
        values = [x for row in matrix for x in row if x is not None]

        # Calcolo i valori richiesti
        if values:
            minimo = min(values)
            massimo = max(values)
            media = sum(values) / len(values)
            mediana = calculate_median(values)
            devStd = calculate_standard_deviation(values, media)
        else:
            minimo = massimo = media = mediana = devStd = None

        simVive = 0
        wsDatiGen = wb[f"Dati Gen{i}"]  # Apro i vari fogli "Dati Gen" per contare il numero di simulazioni vive
        for z in range(2, wsDatiGen.max_row + 1):
            if wsDatiGen.cell(row=z, column=2).value is not None:
                simVive = simVive + 1

        # Scrivo i valori calcolati
        ws.cell(row=2, column=i+1, value=minimo)
        ws.cell(row=3, column=i+1, value=media)
        ws.cell(row=4, column=i+1, value=mediana)
        ws.cell(row=5, column=i+1, value=massimo)
        ws.cell(row=6, column=i+1, value=devStd)
        ws.cell(row=7, column=i+1, value=simVive)
    
def main():
    species, num_generations = read_species_and_generations(ANGULAR_PARAMS)
    if not species:
        print("Nessuna specie trovata nel file di parametri.")
        return
    synthesis_data = load_synthesis_files()
    if not synthesis_data:
        print("Nessun file di sintesi trovato.")
        return
    
    try:
        wb = load_workbook(DISTANZA_ANGOLARE)
    except Exception as e:
        print(f"Errore durante il caricamento del file: {e}")
        return

    create_generation_sheets(wb, synthesis_data, num_generations, species)
    
    # Calcola la distanza angolare per ogni generazione
    for sheet in wb.sheetnames:
        if sheet.startswith("Dati Gen"):
            calculate_angular_distance(wb[sheet], species)

    # Leggi TOTAL_SIM dal file di parametri
    def read_total_simulations(file_path):
        with open(file_path, 'r') as file:
            for line in file:
                if line.startswith("TOTAL_SIM"):
                    _, value = line.strip().split()
                    return int(value)
        return 0

    total_simulations = read_total_simulations(ANGULAR_PARAMS)
    update_and_reorder_distance_matrices(wb, total_simulations)
    update_datiGen(wb, total_simulations)

    # Colora la prima colonna di blu
    for sheet in wb.sheetnames:
        if sheet.startswith("Dati Gen"):
            ws = wb[sheet]
            blue_fill = PatternFill(start_color="4287f5", end_color="4287f5", fill_type="solid")
            for i in range(0, total_simulations + 1):
                ws.cell(row=i + 1, column=1).fill = blue_fill

    sintesiSheet(wb, total_simulations, num_generations)
    print("Total simulations: ", total_simulations)
    print("Num generations: ", num_generations)

    try:
        wb.save(DISTANZA_ANGOLARE)
        print(f"File salvato con successo in {DISTANZA_ANGOLARE}")
    except Exception as e:
        print(f"Errore durante il salvataggio del file: {e}")

if __name__ == "__main__":
    main()
